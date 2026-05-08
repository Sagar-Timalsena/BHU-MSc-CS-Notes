import numpy as np
from openpyxl import Workbook

# -----------------------------
# PARAMETERS
# -----------------------------
num_miners = 300
num_servers = 5
communication_range = 30
area_size = 100

# -----------------------------
# GENERATE POSITIONS
# -----------------------------
miners_pos = np.random.randint(0, area_size, (num_miners, 2))
servers_pos = np.random.randint(0, area_size, (num_servers, 2))

# -----------------------------
# CONNECTIVITY MATRIX
# -----------------------------
connectivity_matrix = np.zeros((num_miners, num_servers), dtype=int)

for i in range(num_miners):
    distances = np.sqrt(
        (servers_pos[:, 0] - miners_pos[i, 0])**2 +
        (servers_pos[:, 1] - miners_pos[i, 1])**2
    )
    connectivity_matrix[i] = (distances <= communication_range).astype(int)

# -----------------------------
# SAVE TO EXCEL
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "Connectivity Matrix"

# Write column headers (servers)
for j in range(num_servers):
    ws.cell(row=1, column=j+2, value=f"Server_{j+1}")

# Write row headers (miners)
for i in range(num_miners):
    ws.cell(row=i+2, column=1, value=f"Miner_{i+1}")

# Write matrix values
for i in range(num_miners):
    for j in range(num_servers):
        ws.cell(row=i+2, column=j+2, value=int(connectivity_matrix[i][j]))

# Save file
wb.save("connectivity_matrix_300x5.xlsx")

print("✅ Excel file saved as connectivity_matrix_300x5.xlsx")